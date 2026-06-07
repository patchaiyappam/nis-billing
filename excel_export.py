"""
Excel Export Module - Export data to .xlsx files using openpyxl.
Files are saved in the exports/ folder.
"""
import os
from datetime import datetime
from config import EXPORTS_DIR

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ── Shared styling helpers ─────────────────────────────────

HEADER_FONT = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF") if OPENPYXL_AVAILABLE else None
HEADER_FILL = PatternFill(start_color="1B2838", end_color="1B2838", fill_type="solid") if OPENPYXL_AVAILABLE else None
BODY_FONT = Font(name="Segoe UI", size=11) if OPENPYXL_AVAILABLE else None
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
) if OPENPYXL_AVAILABLE else None


def _style_header(ws, headers):
    """Write styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _style_row(ws, row_num, values):
    """Write a styled data row."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER


def _auto_width(ws):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def _save_workbook(wb, prefix):
    """Save workbook with timestamped filename. Returns filepath."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{timestamp}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)
    wb.save(filepath)
    return filepath


# ── Export Functions ───────────────────────────────────────

def export_products():
    """Export all products to Excel. Returns filepath or None."""
    if not OPENPYXL_AVAILABLE:
        return None

    from database import get_all_products
    products = get_all_products()

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    _style_header(ws, ["ID", "Product Name", "Price (Rs.)"])
    for i, p in enumerate(products, 2):
        _style_row(ws, i, [p["id"], p["name"], p["price"]])

    _auto_width(ws)
    return _save_workbook(wb, "Products")


def export_customers():
    """Export all customers to Excel. Returns filepath or None."""
    if not OPENPYXL_AVAILABLE:
        return None

    from database import get_all_customers
    customers = get_all_customers()

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    _style_header(ws, ["ID", "Name", "Phone", "Total Due (Rs.)"])
    for i, c in enumerate(customers, 2):
        _style_row(ws, i, [c["id"], c["name"], c["phone"], c["total_due"]])

    _auto_width(ws)
    return _save_workbook(wb, "Customers")


def export_pending():
    """Export only customers with due > 0. Returns filepath or None."""
    if not OPENPYXL_AVAILABLE:
        return None

    from database import get_all_customers
    pending = [c for c in get_all_customers() if c["total_due"] > 0]

    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Dues"

    _style_header(ws, ["ID", "Name", "Phone", "Pending Due (Rs.)"])
    for i, c in enumerate(pending, 2):
        _style_row(ws, i, [c["id"], c["name"], c["phone"], c["total_due"]])

    _auto_width(ws)
    return _save_workbook(wb, "Pending_Dues")
